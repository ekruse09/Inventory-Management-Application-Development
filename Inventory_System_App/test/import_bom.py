
# Import Block
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox

def find_column_index(sheet, header_name):
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value and header_name.lower() in str(cell_value).lower():
            return col
    return None

def process_inventory(inventory_file, bom_file):
    inventory_wb = openpyxl.load_workbook(inventory_file)
    bom_wb = openpyxl.load_workbook(bom_file)

    inventory_ws = inventory_wb.active
    bom_ws = bom_wb.active

    inventory_part_col = find_column_index(inventory_ws, "part number")
    inventory_qty_col = find_column_index(inventory_ws, "qty")
    
    bom_part_col = find_column_index(bom_ws, "part number")
    bom_qty_col = find_column_index(bom_ws, "qty")

    if not all([inventory_part_col, inventory_qty_col, bom_part_col, bom_qty_col]):
        messagebox.showerror("Error", "Could not find the required columns in one or both files.")
        return

    missing_parts = []
    negative_balance_parts = []

    inventory_dict = {}
    for row in range(2, inventory_ws.max_row + 1):  # Skip header
        part_number = inventory_ws.cell(row=row, column=inventory_part_col).value
        qty_on_hand = inventory_ws.cell(row=row, column=inventory_qty_col).value
        inventory_dict[part_number] = {'row': row, 'qty': qty_on_hand}

    for row in range(2, bom_ws.max_row + 1):  # Skip header
        bom_part_number = bom_ws.cell(row=row, column=bom_part_col).value
        bom_qty = bom_ws.cell(row=row, column=bom_qty_col).value

        if bom_part_number in inventory_dict:
            inventory_row = inventory_dict[bom_part_number]['row']
            current_qty = inventory_dict[bom_part_number]['qty']

            if current_qty >= bom_qty:
                inventory_ws.cell(row=inventory_row, column=inventory_qty_col).value = current_qty - bom_qty
            else:
                negative_balance_parts.append((bom_part_number, bom_qty, current_qty))
        else:
            missing_parts.append((bom_part_number, bom_qty))

    if missing_parts or negative_balance_parts:
        root = tk.Tk()
        root.withdraw()

        if missing_parts:
            missing_message = "Missing parts:\n" + "\n".join([f"Part: {part}, Qty: {qty}" for part, qty in missing_parts])
            messagebox.showwarning("Missing Parts", missing_message)

        if negative_balance_parts:
            negative_message = "Negative balance warning:\n" + "\n".join([f"Part: {part}, Required: {req}, On Hand: {on_hand}" for part, req, on_hand in negative_balance_parts])
            messagebox.showwarning("Negative Balance", negative_message)

    inventory_wb.save(inventory_file)
    messagebox.showinfo("Success", "Inventory file has been updated.")

def select_file(entry_widget):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, file_path)

def run_process():
    inventory_file = inventory_entry.get()
    bom_file = bom_entry.get()

    if not inventory_file or not bom_file:
        messagebox.showerror("Error", "Please select both files.")
        return

    process_inventory(inventory_file, bom_file)

# Create the main application window
root = tk.Tk()
root.title("Inventory Processor")

# Inventory file selection
tk.Label(root, text="Select Inventory File:").grid(row=0, column=0, padx=10, pady=10)
inventory_entry = tk.Entry(root, width=50)
inventory_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=lambda: select_file(inventory_entry)).grid(row=0, column=2, padx=10, pady=10)

# BOM file selection
tk.Label(root, text="Select BOM File:").grid(row=1, column=0, padx=10, pady=10)
bom_entry = tk.Entry(root, width=50)
bom_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=lambda: select_file(bom_entry)).grid(row=1, column=2, padx=10, pady=10)

# Run button
tk.Button(root, text="Run", command=run_process).grid(row=2, column=0, columnspan=3, pady=20)

# Start the main event loop
root.mainloop()
